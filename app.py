from flask import Flask, request, render_template, jsonify, send_file, redirect, url_for
import pickle
import numpy as np
from visualization import create_visualization
from pdf_generator import generate_pdf
from email_sender import send_email  # Import the email sending function
import os
import openpyxl

# Load the trained model
model_filename = 'random_forest_model.pkl'
with open(model_filename, 'rb') as file:
    model = pickle.load(file)

# Initialize the Flask app
app = Flask(__name__)

# Define valid ranges for each feature
valid_ranges = {
    'age': (0, 120),
    'gender': (0, 1),
    'chestpain': (0, 3),
    'restingBP': (80, 200),
    'serumcholestrol': (100, 600),
    'fastingbloodsugar': (0, 1),
    'restingrelectro': (0, 2),
    'maxheartrate': (60, 220),
    'exerciseangia': (0, 1),
    'oldpeak': (0, 6.0),
    'slope': (0, 2),
    'noofmajorvessels': (0, 4)
}

# Validate input values
def validate_input(features):
    for feature, value in features.items():
        min_val, max_val = valid_ranges[feature]
        if not (min_val <= value <= max_val):
            return False, f"{feature} is out of range. Expected range: {min_val} to {max_val}."
    return True, ""

# Define the home route
@app.route('/')
def home():
    return render_template('index.html')

# Define the prediction route
@app.route('/predict', methods=['POST'])
def predict():
    # Get the form data
    form_data = request.form
    # Extract the name and phone number (new fields)
    name = form_data.get('name')
    phone = form_data.get('phone')
    
    # Extract feature values from the form data
    features = {
        'age': float(form_data.get('age')),
        'gender': float(form_data.get('gender')),
        'chestpain': float(form_data.get('chestpain')),
        'restingBP': float(form_data.get('restingBP')),
        'serumcholestrol': float(form_data.get('serumcholestrol')),
        'fastingbloodsugar': float(form_data.get('fastingbloodsugar')),
        'restingrelectro': float(form_data.get('restingrelectro')),
        'maxheartrate': float(form_data.get('maxheartrate')),
        'exerciseangia': float(form_data.get('exerciseangia')),
        'oldpeak': float(form_data.get('oldpeak')),
        'slope': float(form_data.get('slope')),
        'noofmajorvessels': float(form_data.get('noofmajorvessels'))
    }
    
    # Validate input values
    is_valid, message = validate_input(features)
    if not is_valid:
        return jsonify({'error': message}), 400
    
    # Convert features to a NumPy array
    feature_values = np.array(list(features.values())).reshape(1, -1)
    
    # Make a prediction for cardiovascular disease
    cvd_prediction = model.predict(feature_values)[0]
    
    # Compute the prediction accuracy (for example, assuming this value is retrieved or calculated)
    accuracy = 98  
    
    # Additional conditions checks
    age, gender, chestpain, restingBP, serumcholestrol, fastingbloodsugar, restingrelectro, maxheartrate, exerciseangia, oldpeak, slope, noofmajorvessels = feature_values[0]
    
    # Conditions for other risk factors diseases
    hypertension = restingBP > 140
    diabetes = fastingbloodsugar > 120
    hyperlipidemia = serumcholestrol > 200
    myocardial_infarction = (chestpain == 1 or chestpain == 2) and restingBP > 140 and serumcholestrol > 200 and exerciseangia == 1 and oldpeak > 1.0 and slope in [1, 2] and noofmajorvessels > 0 and restingrelectro != 0
    
    # Prepare prediction results
    predictions = {
        'Cardiovascular Disease': cvd_prediction,
        'Hypertension': hypertension,
        'Diabetes': diabetes,
        'Hyperlipidemia': hyperlipidemia,
        'Myocardial Infarction': myocardial_infarction
    }
    
    # Generate the visualization and initiate text-to-speech part
    plot_url = create_visualization(predictions)
    
    # Determine if a recommendation should be shown
    show_recommendation = cvd_prediction == 1
    
    # Store user data for PDF generation
    user_data = {
        'name': name,
        'age': age,
        'gender': gender,
        'phone': phone,
        'chestpain': chestpain,
        'restingBP': restingBP,
        'serumcholestrol': serumcholestrol,
        'fastingbloodsugar': fastingbloodsugar,
        'maxheartrate': maxheartrate
    }
    
    # Generate PDF report
    pdf_path = 'static/report.pdf'
    generate_pdf(**user_data, file_path=pdf_path)
    
    # Return the result page with the predictions, plot, and accuracy
    return render_template('result.html', predictions=predictions, plot_url=plot_url, show_recommendation=show_recommendation, accuracy=accuracy, user_data=user_data)

# Define the route to download the PDF report and save user info to Excel
@app.route('/download_report')
def download_report():
    # Path to the Excel file
    excel_filename = 'static/user_data.xlsx'
    
    # Extract user data from the request arguments
    name = request.args.get('name')
    age = request.args.get('age')
    phone = request.args.get('phone')

    # Load the existing workbook or create a new one
    try:
        workbook = openpyxl.load_workbook(excel_filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    
    # Select the active worksheet
    sheet = workbook.active

    # If the file is new and empty, create headers
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet['A1'].value is None:
        sheet.append(['Name', 'Age', 'Phone'])

    # Append the new user data as a row in the Excel sheet
    sheet.append([name, age, phone])

    # Save the updated Excel file
    workbook.save(excel_filename)

    # Send the PDF file for download
    return send_file('static/report.pdf', as_attachment=True)

# Define the recommendation route
@app.route('/recommendation')
def recommendation():
    recommendations = [
        "Maintain a healthy diet rich in fruits, vegetables, and whole grains.",
        "Engage in regular physical activity, such as walking, jogging, or swimming, at least 30 minutes a day.",
        "Avoid smoking and limit alcohol consumption.",
        "Manage stress through relaxation techniques like yoga, meditation, or deep breathing exercises.",
        "Monitor your blood pressure, cholesterol, and blood sugar levels regularly.",
        "Consult your healthcare provider for regular check-ups and follow their advice regarding medications or lifestyle changes."
    ]
    return render_template('recommendation.html', recommendations=recommendations)

# Define the recommendations text route
@app.route('/recommendations_text')
def recommendations_text():
    recommendations = [
        "Maintain a healthy diet rich in fruits, vegetables, and whole grains.",
        "Engage in regular physical activity, such as walking, jogging, or swimming, at least 30 minutes a day.",
        "Avoid smoking and limit alcohol consumption.",
        "Manage stress through relaxation techniques like yoga, meditation, or deep breathing exercises.",
        "Monitor your blood pressure, cholesterol, and blood sugar levels regularly.",
        "Consult your healthcare provider for regular check-ups and follow their advice regarding medications or lifestyle changes."
    ]
    recommendations_text = " ".join(recommendations)
    return jsonify({'text': recommendations_text})

# Define the route to send the PDF report via email
@app.route('/send_email', methods=['POST'])
def send_email_route():
    # Get the email address from the form data
    email = request.form.get('email')
    
    # Path to the PDF file
    pdf_path = 'static/report.pdf'
    
    # Send the email
    result = send_email(email, pdf_path)
    
    # Check if the email was sent successfully
    if 'successfully' in result:
        return jsonify({'success': result})
    else:
        return jsonify({'error': result}), 500

# Define the route to render the email form
@app.route('/email_form')
def email_form():
    return render_template('email_form.html')

if __name__ == '__main__':
    app.run(debug=True)
